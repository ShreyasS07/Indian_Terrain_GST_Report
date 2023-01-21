import os
import openpyxl
import pandas as pd
from tkinter import *
from tkinter import filedialog

def select_folder():
    global folder_path
    folder_path = filedialog.askdirectory()
    print("Selected folder path is : ", folder_path)
    # files = os.listdir(folder_path)

    # Get list of files in selected folder
    file_list = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    print('Files inside the selected folder is:')
    for file in file_list:
        print(file)

    # # Creating separate dataframes for csv and xlsx files
    # csv_df = pd.DataFrame()
    # xlsx_df = pd.DataFrame()
    # files = os.listdir(folder_path)
    # for file in files:
    #     if file.endswith('.csv'):
    #         df = pd.read_csv(os.path.join(folder_path, file))
    #         print("Dataframe shape for", file, ":", df.shape)
    #     elif file.endswith('.xlsx'):
    #         df = pd.read_excel(os.path.join(folder_path, file))
    #         print("Dataframe shape for", file, "is :", df.shape)
    # # Combining all dataframes into single dataframe
    # final_df = pd.concat([csv_df, xlsx_df])
    # print("Shape of final dataframe:", final_df.shape)

    # df1 = pd.read_excel(r"C:\Users\ASUS\Downloads\Nov GST Set\Nov GST Set\Sales 08th to 10th Nov-22.xlsb.xlsx")
    # df1.head()
    # df1.isnull().sum()

    # Create dataframes for each CSV and XLSX file
    for file in file_list:
        if file.endswith(".csv"):
            df = pd.read_csv(os.path.join(folder_path, file))
            df_name = file.replace(".csv", "") + "_df"
            globals()[df_name] = df
            print(df_name, "shape:", df.shape)
        elif file.endswith(".xlsx"):
            df = pd.read_excel(os.path.join(folder_path, file))
            df_name = file.replace(".xlsx", "") + "_df"
            globals()[df_name] = df
            print(df_name, "shape:", df.shape)

    # Combine all dataframes into a single dataframe
    final_df = pd.concat([v for k, v in globals().items() if k.endswith("_df")], ignore_index=True)

    print("final_df shape is :", final_df.shape)
    print("Column Names:", final_df.columns)
    print("\nData Count:")
    print(final_df.count())
    print("Final dataframe contains null values:", final_df.isnull().values.any())
    print("Final dataframe contains duplicate data:", final_df.duplicated().any())

    #Saving the final Dataframe into Excel file in the same folder

    final_df.to_excel(os.path.join(folder_path, "Final report.xlsx"), index=False)
    print("All Excel files extracted & saved to single excel file & stored in same folder")
    print("Final report saved at:", os.path.join(folder_path, "Final report.xlsx"))

    # #Saving final_df to xlsb format.
    # final_df.to_excel(os.path.join(folder_path, "Final report.xlsb"), index=False)
    #
    # DF_Output = pd.read_excel(os.path.join(folder_path, "Final report.xlsx"))
    # DF_Output.shape()






    # ------------------------------------------   xlsx to xlsb  ----------------------------------------------

    # Reducing the final report size by Converting into Binary Excel format.

    # workbook_xlsx = openpyxl.load_workbook(folder_path, "Final report.xlsx")
    # print("final report loaded ")
    #
    # # Create a new xlsb file
    # workbook_xlsb = openpyxl.Workbook(write_only=True)
    #
    # # Iterate over the worksheets in the xlsx file
    # for sheet in workbook_xlsx:
    #     # Create a new worksheet in the xlsb file
    #     ws_xlsb = workbook_xlsb.create_sheet(sheet.title)
    #     # Copy the data from the xlsx worksheet to the xlsb worksheet
    #     for row in sheet.iter_rows():
    #         ws_xlsb.append([cell.value for cell in row])
    #
    # # Save the xlsb file
    # workbook_xlsb.save(os.path.join(folder_path, "Final Binary report.xlsb"))
    # print("Final Binary report created ")



    # ------------------------------------------   Creating Pivot table  ----------------------------------------------

    # # Creating Pivot table
    # print("Creating a Pivot table for Final report")
    #
    # # taking created Excel file from previous code
    # file_path = 'data/base_file.xlsx'
    # df = pd.read_excel(file_path)
    #
    # # Create a pivot table
    # pivot_table = df.pivot_table(values='Value', index='Category', columns='Year', aggfunc='sum')
    #
    # # Save the pivot table to a new Excel file
    # pivot_file_path = 'data/pivot_file.xlsx'
    # pivot_table.to_excel(pivot_file_path)


window = Tk()
window.title(" Mindful Automation Pvt Ltd ")
window.geometry('600x150')
window.configure(background="white")
label_file_explorer = Label(window, text=" Indian Terrain GST Report Automation ",
                            width=100, height=3, fg="blue")
file = Button(window,
                  text=" Select the folder ", command=select_folder)
label_file_explorer.grid(column=1, row=1)
file.grid(column=1, row=2)
window.mainloop()